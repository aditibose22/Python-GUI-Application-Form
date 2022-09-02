#CSE 2nd Year 3rd Sem Roll no. 5,8,9,10,12
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import pathlib
import openpyxl, xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from win32api import GetSystemMetrics
f1 = Tk()
w=GetSystemMetrics(0)
h=GetSystemMetrics(1)
f=StringVar()
f="%dx%d" %(w,h)
f1.geometry('950x700')
f1.minsize(950,700)
# f1.geometry(f)

ff1=LabelFrame(f1,height=0.16*h,width=0.9*w,padx=5,pady=3)
ff1.pack(padx=5,pady=3)

ff2=LabelFrame(f1,height=0.310*h,width=0.9*w,padx=5,)
ff2.pack()

ff3=LabelFrame(f1,height=0.325*h,width=0.9*w,padx=5,pady=2)
ff3.pack(padx=5,pady=5)

ff4=LabelFrame(f1,height=36,width=1235,padx=5,pady=1)
ff4.pack(padx=5,pady=1)


f1.title("Student Information")
f1.iconbitmap(r'logo.ico')
img=tk.PhotoImage(file="m_OIP.png")
l = Label(ff1, text=" ST.THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font=('Arial', 16), fg="blue", bg="white")
l.place(x=170,y=1)
l["compound"]=tk.LEFT
l["image"]=img


l1 = Label(ff2, text='STUDENT INFORMATION', font=('Arial', 15, 'bold'), fg="black", bg="white")
l1.place(x=400,y=3)

file = pathlib.Path('Student_Data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Student's Name"
    sheet["B1"] = "Father's Name"
    sheet["C1"] = "Mother's Name"
    sheet["D1"] = "DOB"
    sheet["E1"] = "Gender"
    sheet["F1"] = "EMAIL ID"
    sheet["G1"] = "STUDENT CONTACT No."
    sheet["H1"] = "GUARDIAN CONTACT NO."
    sheet["I1"] = "CURRENT ADD."
    sheet["J1"] = "PERMANENT ADD."
    sheet["K1"] = "10TH BOARD"
    sheet["L1"] = "10TH MEDIUM"
    sheet["M1"] = "10TH SCORE"
    sheet["N1"] = "12TH BOARD"
    sheet["O1"] = "12TH MEDIUM"
    sheet["P1"] = "12TH %"
    file.save('Student_Data.xlsx')

def isvalidemail(useremail):
   
    if(len(useremail)>7):
        if re.match('^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$',useremail)!=None:
            return True
        else:
             messagebox.showwarning('Error!','Invalid email!')
             return False
    else:
        messagebox.showwarning('Error!','Invalid email!')
        return False

def validate_phone(input):
    if input.isdigit():
        return True
    elif input == "":
        return True
    else:
        messagebox.showwarning('Error!', 'Only Digits are allowed!')
        return False
a1 = StringVar()
b4 = StringVar()
c4 = StringVar()
d4 = StringVar()
e4 = StringVar()
f4 = StringVar()
g4 = StringVar()
h4 = StringVar()
i4 = StringVar()
j4 = StringVar()
k4 = StringVar()

stunm = Label(ff2, text="STUDENT'S NAME")
stunm.place(x=50,y=35)
snm = Entry(ff2,textvariable=a1,border=1)
snm.place(x=300,y=35)

fanm = Label(ff2, text="FATHER'S NAME")
fanm.place(x=50,y=55)
fnm = Entry(ff2,textvariable=b4,border=1)
fnm.place(x=300,y=55)

monm = Label(ff2, text="MOTHER'S NAME")
monm.place(x=50,y=75)
mnm = Entry(ff2, textvariable=c4,border=1)
mnm.place(x=300,y=75)

day = [i for i in range(1,32)]
month = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUNE', 'JULY', 'AUGUST', 'SEPT', 'OCT', 'NOV', 'DEC']
year = [i for i in range(1999, 2009)]
s = ['Percentage', 'CGPA']

dob = Label(ff2, text='DATE OF BIRTH')
dob.place(x=50,y=95)

m1 = ttk.Combobox(ff2,textvariable=d4,value=day, state='readonly')
m1.current(0)
m1.place(x=300,y=95)
m1.bind("<<ComboboxSelected>>")

m2 = ttk.Combobox(ff2,textvariable=e4,value=month, state='readonly')
m2.current(0)
m2.place(x=500,y=95)
m2.bind("<<ComboboxSelected>>")

m3 = ttk.Combobox(ff2,textvariable=f4,value=year, state='readonly')
m3.current(0)
m3.place(x=700,y=95)
m3.bind("<<ComboboxSelected>>")

gender = Label(ff2, text='GENDER')
gender.place(x=50,y=116)
gn = IntVar()
Radiobutton(ff2, text='Male', variable=gn, value=1).place(x=300,y=116)
Radiobutton(ff2, text='Female', variable=gn, value=2).place(x=500,y=116)
Radiobutton(ff2, text='Other', variable=gn, value=3).place(x=700,y=116)

em = Label(ff2, text="EMAIL ID")
em.place(x=50,y=133)
emel = Entry(ff2, textvariable=g4,border=1)
emel.place(x=300,y=133)

cno = Label(ff2, text="CONTACT NO. OF STUDENT")
cno.place(x=50,y=153)
cnum = Entry(ff2,textvariable=h4 ,border=1)
cnum.place(x=300,y=153)
vno = ff2.register(validate_phone)
cnum.config(validate="key", validatecommand=(vno, '%P'))

gno = Label(ff2, text="CONTACT NO. OF GUARDIAN")
gno.place(x=50,y=172)
gnum = Entry(ff2,textvariable=i4, border=1)
gnum.place(x=300,y=172)
vno = ff2.register(validate_phone)
gnum.config(validate="key", validatecommand=(vno, '%P'))

curr = Label(ff2, text="CURRENT ADDRESS")
curr.place(x=50,y=191)
cad = Entry(ff2,textvariable=j4, border=1)
cad.place(x=300,y=191)

per = Label(ff2, text="PERMANENT ADDRESS")
per.place(x=50,y=210)
pad = Entry(ff2,textvariable=k4, border=1)
pad.place(x=300,y=210)

b = StringVar()
med = StringVar()
b1 = StringVar()
med1 = StringVar()
po = StringVar()
p = StringVar()
c = StringVar()
ma = StringVar()
s1 = StringVar()
ss1 = StringVar()
s2 = StringVar()
ss2 = StringVar()
per = StringVar()
a=StringVar()
def valid10():
    try:
        if m.get() == 'CGPA':
            t = float(po.get())
            if (t > 10.0 or t < 0.0):
                messagebox.showwarning("Error!", "Enter Valid Input")
            else:
                messagebox.showinfo("Information", "Marks entered successfully!")
        elif m.get() == 'Percentage':
            t = float(po.get())
            if (t > 100.0 or t < 0.0):
                messagebox.showwarning("Error!", "Enter Valid Input")
            else:
                messagebox.showinfo("Information", "Marks entered successfully!")
    except:
        messagebox.showwarning("Error!", "Enter Valid Input")
    

e1 = Label(ff3,text="CLASS 10", bg="grey", width=30, font=('Arial', 12, 'bold')).place(x=370,y=2)
board = Label(ff3,text="BOARD").place(x=50,y=30)
board1 = Entry(ff3, textvariable=b)
board1.place(x=300,y=30)
medium = Label(ff3,text="MEDIUM").place(x=500,y=27)
medium1 = Entry(ff3, textvariable=med)
medium1.place(x=700,y=27)
m = ttk.Combobox(ff3, value=s,textvariable=a,state='readonly')
m.current(0)
m.place(x=500,y=45)
m.bind("<<ComboboxSelected>>")
pp = Entry(ff3, textvariable=po)
pp.place(x=700,y=45)
pp1 = Button(ff3, text="ENTER", command=valid10, font=('bold', 9)).place(x=850,y=45)

e1 = Label(ff3,text="CLASS 12", bg="grey", width=30, font=('Arial', 12, 'bold')).place(x=370,y=75)
board2 = Label(ff3,text="BOARD").place(x=50,y=105)
board3 = Entry(ff3, textvariable=b1)
board3.place(x=300,y=105)
medium2 = Label(ff3,text="MEDIUM").place(x=500,y=105)
medium3 = Entry(ff3, textvariable=med1)
medium3.place(x=700,y=105)
e2 = Label(ff3,text="MARKS OBTAINED [BEST OF 5]", bg="grey", fg='white', width=30, font=('Arial', 10, 'bold')).place(x=400,y=132)
physics = Label(ff3,text="PHYSICS").place(x=50,y=160)
phys1 = Entry(ff3, textvariable=p).place(x=300,y=160)
chem = Label(ff3,text="CHEMISTRY").place(x=500,y=160)
chem1 = Entry(ff3, textvariable=c).place(x=700,y=160)
maths = Label(ff3,text="MATHS").place(x=50,y=180)
maths1 = Entry(ff3, textvariable=ma).place(x=300,y=180)


def picksub(e):
    d = m11.get()
    md = []
    for i in sub:
        if i != d:
            md.append(i)
    m12.config(value=md)
    m12.current(1)


def percentage():
    try:
        f1 = float(p.get())
        f2 = float(c.get())
        f3 = float(ma.get())
        f4 = float(s1.get())
        f5 = float(s2.get())
        if(0.0<=f1<=100.0 and 0.0<=f2<=100.0 and 0.0<=f3<=100.0 and 0.0<=f4<=100.0 and 0.0<=f5<=100.0):
            r = (f1 + f2 + f4 + f5 + f3) / 5
            per.set(r)
        else:
            messagebox.showwarning("Error!", "Enter Valid Input")
    except:
        messagebox.showwarning("Error!", "Enter Valid Input")


sub = ["BIOLOGY", "COMPUTER SCIENCE", "ENGLISH", "BENGALI", "PHYSICAL EDUCATION", "HINDI", "ECONOMICS"]
m11 = ttk.Combobox(ff3, value=sub, textvariable=ss1, state='readonly')
m11.current(0)
m11.place(x=500,y=180)
m11.bind("<<ComboboxSelected>>", picksub)
drop1 = Entry(ff3, textvariable=s1)
drop1.place(x=700,y=180)

m12 = ttk.Combobox(ff3, value=[" "], textvariable=ss2, state='readonly')
m12.current(0)
m12.place(x=50,y=201)
drop2 = Entry(ff3, textvariable=s2)
drop2.place(x=300,y=201)

l = Label(ff3, text="PERCENTAGE")
l.place(x=50,y=222)
l1 = Entry(ff3, textvariable=per)
l1.place(x=300,y=222)
l1 = Button(ff3, text="CALCULATE", command=percentage, font=('bold', 9)).place(x=700,y=215)

def submit():

    a = snm.get()
    b = fnm.get()
    c = mnm.get()
    z = emel.get()
    w = cnum.get()
    v = gnum.get()
    day = m1.get()
    month = m2.get()
    year = m3.get()
    dob = f'{day}/{month}/{year}'
    h = cad.get()
    i = pad.get()
    gen = gn.get()

    k = board1.get()
    l = medium1.get()
    p = pp.get() + m.get()
    n = board3.get()
    r = medium3.get()
    s = per.get()

    if (a == "" or b == "" or c == "" or z == "" or w == "" or v=="" or day == "" or month == "" or year == "" or gen=="" or h == "" or i == "" or dob == "" or k=="" or l=="" or p=="" or n=="" or r=="" or s==""):
        print("error")
        messagebox.showwarning("Error!", "Enter all fields")
    elif (len(w) != 10 or len(v)!=10):
        messagebox.showwarning("Error!", "Enter 10 digit number")
    elif(isvalidemail(z)):  
    
        file = openpyxl.load_workbook('Student_Data.xlsx')
        sheet = file.active
        if gn.get() == 1:
            gen = 'Male'
        elif gn.get() == 2:
            gen = 'Female'
        elif gn.get() == 3:
            gen = 'Other'
        sheet.cell(column=1, row=sheet.max_row + 1, value=a)
        sheet.cell(column=2, row=sheet.max_row, value=b)
        sheet.cell(column=3, row=sheet.max_row, value=c)
        sheet.cell(column=4, row=sheet.max_row, value=dob)
        sheet.cell(column=5, row=sheet.max_row, value=gen)
        sheet.cell(column=6, row=sheet.max_row, value=z)              
              
        
        sheet.cell(column=7, row=sheet.max_row, value=w)
        sheet.cell(column=8, row=sheet.max_row, value=v)
        sheet.cell(column=9, row=sheet.max_row, value=h)
        sheet.cell(column=10, row=sheet.max_row, value=i)
        sheet.cell(column=11, row=sheet.max_row, value=k)
        sheet.cell(column=12, row=sheet.max_row, value=l)
        sheet.cell(column=13, row=sheet.max_row, value=p)
        sheet.cell(column=14, row=sheet.max_row, value=n)
        sheet.cell(column=15, row=sheet.max_row, value=r)
        sheet.cell(column=16, row=sheet.max_row, value=s)
        file.save("Student_Data.xlsx")
        data_xls = pd.read_excel('Student_Data.xlsx', 'Sheet', dtype=str, index_col=None)
        data_xls.to_csv('Student_Data.csv', encoding='utf-8', index=False)
        messagebox.showinfo('Finished!',"Data entered successfully!")
def clear():
    a1.set("")
    b4.set("")
    c4.set("")
    d4.set("")
    e4.set("")
    f4.set("")
    g4.set("")
    h4.set("")
    i4.set("")
    j4.set("")
    k4.set("")
    a.set("")
    b.set("")
    med.set("")
    b1.set("")
    med1.set("")
    po.set("")
    p.set("")
    c.set("")
    ma.set("")
    s1.set("")
    ss1.set("")
    s2.set("")
    ss2.set("")
    per.set("")

submitA = Button(ff4,text="CLEAR",font=('bold',10),command=clear).place(x=50,y=1)
submitB = Button(ff4,text="SUBMIT", font=('bold', 10), command=submit).place(x=400,y=1)
ex = Button(ff4, text='EXIT', fg='blue', font=('bold', 10), command=f1.destroy)
ex.place(x=700,y=1)

f1.mainloop()
